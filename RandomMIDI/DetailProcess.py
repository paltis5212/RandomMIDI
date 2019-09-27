import random
import math
from enum import Enum
from collections import Counter
from openpyxl import load_workbook
from midiutil import MIDIFile
from datetime import datetime


class NoteNum(Enum):
    C = 72
    D = 74
    E = 76
    F = 79
    G = 81
    A = 84
    B = 86
    lifting = 12
    order = ["C", "D", "E", "F", "G", "A", "B"]


TempoNum = {"__": 0.25, "_": 0.5, "*": 1, "-": 2, "--": 3, "---": 4}


class WeightsObject():
    def __init__(self, data: dict):
        # 原始資料
        self.data: dict = data
        self.keyList = list(data.keys())
        self.valuesList = list(data.values())
        # sum、maxList 都 * 100 過了，可用於小數點計算
        # 所有 value 總和
        self.sum: int = 0
        # 每個元素的 max 範圍
        self.maxList = []
        # 計算要用的東西
        for x in self.valuesList:
            self.sum += math.floor(float(x) * 100)
            self.maxList.append(self.sum)

    # 根據權重取得一串隨機結果
    # num: int = 需要幾個結果
    def getRandListKey(self, num: int = 50):
        randList = [random.randint(1, self.sum) for i in range(num)]
        result: list = []
        for rand in randList:
            result.append(self.getNumInRange(rand))
        return result

    # 根據權重取得一個隨機結果
    def getRandKey(self):
        rand = random.randint(1, self.sum)
        return self.getNumInRange(rand)

    # 檢測隨機數在哪個範圍
    # num: int = 隨機數
    def getNumInRange(self, num: int):
        for i in range(len(self.maxList)):
            if num <= self.maxList[i]:
                return self.keyList[i]


class ComputeUtils():
    # 兩個權重結果比較，有順序性
    def weightsCompare(self, list0: list, list1: list):
        # 狀況一
        for i in range(len(list0)):
            # 比較項不存在
            if len(list0[i:]) == 0 or len(list1[i:]) == 0:
                break
            # 兩個隨機數一樣，回傳
            if list0[i] == list1[i]:
                return list0[i]

        # 狀況二
        list0MaxKeys = self.getMaxValueKeys(list0)
        list1MaxKeys = self.getMaxValueKeys(list1)
        # 同一個 key 在兩個列表中出現最多，則取用
        for key in list0MaxKeys:
            if key in list1MaxKeys:
                return key

        # 狀況三
        # 回傳在 list0 中出現最多次的 key
        return random.choice(list0MaxKeys)

    # 取得列表中最大值的每個 key
    # list: list = 要用的 list
    def getMaxValueKeys(self, list: list):
        keyCountDict = Counter(list)
        maxValue = 0
        maxList = []
        for key, value in keyCountDict.items():
            # 大於最大值，代替他
            if value > maxValue:
                maxList = [key]
                maxValue = value
            # 同最大值，放到列表裡
            elif value == maxValue:
                maxList.append(key)
        return maxList

    # 使用升降幾個音來取得需要的音符
    # passNoteStr: str = 上一個音符
    # step: int = 要升降的幅度
    def stepToScale(self, passNoteStr: str, step: int):
        passNote = passNoteStr[: 1]
        newNote = int(passNote) + step
        newNoteStr = passNoteStr
        # 重複升降音階
        while newNote < 1 or newNote > 7:
            # 降音階
            if newNote < 1:
                newNote += 7
                newNoteStr = self.cutOrAddNoteLifting(newNoteStr, ",")
            # 升音階
            if newNote > 7:
                newNote -= 7
                newNoteStr = self.cutOrAddNoteLifting(newNoteStr, "'")
        # 改音符
        result = 7 if newNote % 7 == 0 else newNote
        newNoteStr = str(result) + newNoteStr[1:]
        return newNoteStr

    # 減掉或新增後面的升降音階符號
    # targetStr: str = 目標音符
    # lifting: str = 需要升或降
    def cutOrAddNoteLifting(self, targetStr: str, lifting: str):
        # 最後的升降和目標升降一樣，或後面沒有符號，則後面加一個
        if targetStr[-1] == lifting or len(targetStr) == 1:
            targetStr += lifting
        # 最後的升降和目標升降不一樣，後面減一個
        elif targetStr[-1] != lifting:
            targetStr = targetStr[: -1]
        return targetStr

    # 給出音符文字代表的數字(根據 midi)
    def noteStrToNum(self, noteStr):
        up: int = noteStr.count("'")
        down: int = noteStr.count(",")
        note: int = noteStr[: 1]
        order: list = NoteNum.order.value
        lifting: int = NoteNum.lifting.value
        return NoteNum[order[int(note) - 1]].value + (up - down) * lifting

    def getScaleAndTempo(self, noteStr: str):
        """ 完整音符文字中，取得音階和節拍 """
        index = 0
        for strIndex in range(len(noteStr)):
            str = noteStr[strIndex]
            if str.isdigit() or str == "'" or str == ",":
                index = strIndex + 1
        return {
            "scale": noteStr[: index],
            "tempo": noteStr[index:]
        }


class ExcelMgr():
    def getAllTable(self, path: str):
        """ 得到精美設計後的表單內容 """
        wb = load_workbook(path)
        ws = wb.active
        titleRow = {}
        nowTable = "limit"
        try:
            # 獲得標題 row
            for cell in ws["A"]:
                if cell.value == "limit" or cell.value == "weights":
                    nowTable = cell.value
                    titleRow[nowTable] = {}
                elif cell.value != None:
                    titleRow[nowTable][cell.value] = cell.row

            allTable = {}
            # 讀取 row 資料
            for table in titleRow:
                allTable[table] = {}
                for title in titleRow[table]:
                    # 讀取 row 在 sheet 資料，給 key
                    allTable[table][title] = {}
                    for x in ws[titleRow[table][title]]:
                        key = ws.cell(row=x.row - 1, column=x.column).value
                        if x.value != None and key != None:
                            allTable[table][title][key] = x.value
        except:
            print("ERROR: ExcelMgr.getAllTable")
        return allTable


class MidiMgr():
    def __init__(self, music: dict, bpm: int, program: int):
        MyMIDI = MIDIFile(1)
        track = 0   # the only track
        time = 0    # start at the beginning
        MyMIDI.addTrackName(track, time, "simple")
        MyMIDI.addTempo(track, time, int(bpm))
        MyMIDI.addProgramChange(0, 0, 0, int(program))
        # add some notes
        channel = 0
        for sentences in music:
            for note in music[sentences]:
                if note == "0":
                    time += 1
                    continue
                scaleAndTempo = ComputeUtils().getScaleAndTempo(note)
                scaleNum = ComputeUtils().noteStrToNum(scaleAndTempo["scale"])
                tempoNum = TempoNum[scaleAndTempo["tempo"]]
                MyMIDI.addNote(track=track, channel=channel, pitch=scaleNum,
                               time=time, duration=tempoNum, volume=random.choice([75, 100, 120]))
                time += tempoNum
        with open(datetime.now().strftime('%Y%m%d-%H%M%S') + ".mid", "wb") as midi_file:
            MyMIDI.writeFile(midi_file)
