from openpyxl import load_workbook
from Enums import ScalePitch, LimitData, WeightsData
from InfoMgr import InfoMgr
from ObjectModel import WeightsObject, SentenceObject
from enum import Enum


class Main():
    def __init__(self):
        # 要先做，取得表單內容後比較好做
        self._setTableData("rules.xlsx")
        self._setMinAndMax()
        self.repeatNum = 0
        for sentenceIndex in range(int(InfoMgr.limit[LimitData.sentenceCount.value]["value"])):
            # 重複
            if self._repeatSentence(sentenceIndex) or self.repeatNum in InfoMgr.music:
                continue
            # 新句子
            sentenceObject = SentenceObject(pos=sentenceIndex)
            sentenceObject.setNew()
            InfoMgr.music[sentenceIndex] = sentenceObject
            self.repeatNum += 1

    def _setTableData(self, path: str):
        """ 設定精美設計後的表單內容到 limit 和 weights 
        :param path: excel 檔路徑
        """
        wb = load_workbook(path)
        ws = wb.active
        titleRow = {}
        nowTable = LimitData.limit.value
        try:
            # 獲得標題 row
            for cell in ws["A"]:
                if cell.value == LimitData.limit.value or cell.value == WeightsData.weights.value:
                    nowTable = cell.value
                    titleRow[nowTable] = {}
                elif cell.value != None:
                    titleRow[nowTable][cell.value] = cell.row
            allTable = {}
            # 讀取 row 資料
            for table in titleRow:
                allTable[table] = {}
                for title in titleRow[table]:
                    # 讀取 row 在 sheet 資料，給 key
                    allTable[table][title] = {}
                    for x in ws[titleRow[table][title]]:
                        key = ws.cell(row=x.row - 1, column=x.column).value
                        if x.value != None and key != None:
                            allTable[table][title][key] = x.value
            # 設定到 InfoMgr
            InfoMgr.limit = allTable[LimitData.limit.value]
            InfoMgr.weights = allTable[WeightsData.weights.value]
            self._setWeightsObject()
        except:
            print("ERROR ", self._setTableData.__name__)

    def _setWeightsObject(self):
        """ 把單純的字典轉成 WeightsObject """
        for key in InfoMgr.weights:
            weightsObject = WeightsObject(InfoMgr.weights[key])
            InfoMgr.weights[key] = weightsObject

    def _setMinAndMax(self):
        """ 設定最大最小的 pitch 範圍 """
        InfoMgr.minPitch = self._noteStrToPitch(
            InfoMgr.limit[LimitData.scale.value]["min"])
        InfoMgr.maxPitch = self._noteStrToPitch(
            InfoMgr.limit[LimitData.scale.value]["max"])

    def _noteStrToPitch(self, noteStr: str):
        """ 純文字音符變成 pitch
        :param noteStr: 純文字音符，例如 7'
        """
        up: int = noteStr.count("'")
        down: int = noteStr.count(",")
        note: int = noteStr[: 1]
        return ScalePitch[ScalePitch.scaleName.value[int(note) - 1]].value + (up - down) * ScalePitch.step.value

    def _repeatSentence(self, sentenceIndex: int):
        repeatCount = int(
            InfoMgr.weights[WeightsData.repeatCount.value].getRandKey())
        # 沒有那麼多句子可以重複，不給過
        if repeatCount > self.repeatNum:
            return False
        # 重複很多句子
        else:
            for index in range(repeatCount):
                nowIndex = sentenceIndex + index
                sentenceObject = SentenceObject(pos=nowIndex)
                sentenceObject.repeatFrom(
                    InfoMgr.music[nowIndex - repeatCount])
                InfoMgr.music[nowIndex] = sentenceObject
            self.repeatNum = 0
            return True


a = Main()
