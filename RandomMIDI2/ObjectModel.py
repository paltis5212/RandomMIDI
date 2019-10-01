import random
from openpyxl import load_workbook
from ComputeUtils import ComputeUtils
from Enums import *


class Weights():
    def __init__(self, data: dict):
        self.keyList = list(data.keys())
        self.valueList = list(data.values())

    def getRandKey(self):
        """ 獲得一個隨機值 """
        return random.choices(population=self.keyList, weights=self.valueList)[0]


class Note():
    track: int
    channel: int
    pitch: int
    duration: float
    volume: int
    scale: int
    scaleRaise: int


class Sentence():
    startWordCount: int
    endWordCount: int
    wordCount: int
    pos: int
    repeatFrom: int
    notes: list

    def __init__(self, repeatFrom=None, wordWeights: Weights = None, endWeights: Weights = None):
        isInit = False
        # 指定重複句子
        if repeatFrom is not None:
            self.startWordCount = repeatFrom.startWordCount
            self.endWordCount = repeatFrom.endWordCount
            self.wordCount = repeatFrom.wordCount
            self.repeatFrom = repeatFrom.pos
            self.notes = repeatFrom.notes
        # 新語句生成字數規則
        elif wordWeights is not None and endWeights is not None:
            while self.wordCount < self.endWordCount or not isInit:
                isInit = True
                self.wordCount = int(wordWeights.getRandKey())
                self.endCount = int(endWeights.getRandKey())
            self.startWordCount = self.wordCount - self.endWordCount
        else:
            print("ERROR Sentence.__init__")

    def randomChangeNotes(self, changeCount: int, music: dict, sentence: Sentence, index: int, scaleFluctuationWeight: Weights, scaleWeight: Weights, tempoWeight: Weights, strengthWeight: Weights):
        """ 改變指定數量的音符 """
        changeIndex = []
        index = 0
        # 不能超過句子音符數(外面要先過濾一次，通常不希望這條成立)
        if changeCount > self.wordCount:
            changeCount = self.wordCount
        # 生成要改變的音符編號(不重複)
        while index < changeCount:
            num = random.randint(0, self.wordCount)
            if num not in changeIndex:
                changeIndex.append(num)
                index += 1
        # 生成新音符
        for index in changeIndex:
            self.notes[index] = ComputeUtils().createNewNoteFromPass(
                music, sentence, index, scaleFluctuationWeight, scaleWeight, tempoWeight, strengthWeight)

    def AllTable(self):
        def __init__(self, path: str):
            wb = load_workbook(path)
            ws = wb.active
            titleRow = {}
            nowTable = LimitData.limit
            try:
                # 獲得標題 row
                for cell in ws["A"]:
                    if cell.value == LimitData.limit or cell.value == WeightsData.weights:
                        nowTable = cell.value
                        titleRow[nowTable] = {}
                    elif cell.value != None:
                        titleRow[nowTable][cell.value] = cell.row

                allTable = {}
                # 讀取 row 資料
                for table in titleRow:
                    allTable[table] = {}
                    for title in titleRow[table]:
                        # 讀取 row 在 sheet 資料，給 key
                        allTable[table][title] = {}
                        for x in ws[titleRow[table][title]]:
                            key = ws.cell(row=x.row - 1, column=x.column).value
                            if x.value != None and key != None:
                                allTable[table][title][key] = x.value
                self.limit = allTable[LimitData.limit]
                self.weights = allTable[WeightsData.weights]
            except:
                print("ERROR AllTable.__init__")
