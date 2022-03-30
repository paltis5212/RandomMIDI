#! -*- coding:utf-8 -*-
import sys
sys.path.insert(0, "/Users/user/.pyenv/versions/3.6.8/lib/python3.6/site-packages")
from openpyxl import load_workbook
from Enums import ScalePitch, LimitData, WeightsData
from InfoMgr import InfoMgr
from ObjectModel import WeightsObject, SentenceObject
from enum import Enum
from midiutil import MIDIFile
from datetime import datetime


class Main():
    def __init__(self):
        # 要先做，取得表單內容後比較好做
        self._setTableData("RandomMIDI3/rules.xlsx")
        self._setMinAndMax()
        self.repeatNum = 0
        for sentenceIndex in range(int(InfoMgr.limit[LimitData.sentenceCount.value]["value"])):
            # 已經有句子了
            if sentenceIndex in InfoMgr.music:
                continue
            repeatResult = self._repeatSentence(sentenceIndex)
            # 本次已複製句子
            if repeatResult:
                continue
            # 新句子
            sentenceObject = SentenceObject(pos=sentenceIndex)
            sentenceObject.setNew()
            InfoMgr.music[sentenceIndex] = sentenceObject
            self.repeatNum += 1
        self._createMIDI()

    def _setTableData(self, path: str):
        """ 設定精美設計後的表單內容到 limit 和 weights 
        :param path: excel 檔路徑
        """
        wb = load_workbook(path)
        ws = wb.active
        titleRow = {}
        nowTable = LimitData.limit.value
        try:
            # 獲得標題 row
            for cell in ws["A"]:
                if cell.value == LimitData.limit.value or cell.value == WeightsData.weights.value:
                    nowTable = cell.value
                    titleRow[nowTable] = {}
                elif cell.value != None:
                    titleRow[nowTable][cell.value] = cell.row
            allTable = {}
            # 讀取 row 資料
            for table in titleRow:
                allTable[table] = {}
                for title in titleRow[table]:
                    # 讀取 row 在 sheet 資料，給 key
                    allTable[table][title] = {}
                    for x in ws[titleRow[table][title]]:
                        key = ws.cell(row=x.row - 1, column=x.column).value
                        if x.value != None and key != None:
                            allTable[table][title][key] = x.value
            # 設定到 InfoMgr
            InfoMgr.limit = allTable[LimitData.limit.value]
            InfoMgr.weights = allTable[WeightsData.weights.value]
            self._setWeightsObject()
        except:
            print("ERROR ", self._setTableData.__name__)

    def _setWeightsObject(self):
        """ 把單純的字典轉成 WeightsObject """
        for key in InfoMgr.weights:
            weightsObject = WeightsObject(InfoMgr.weights[key])
            InfoMgr.weights[key] = weightsObject

    def _setMinAndMax(self):
        """ 設定最大最小的 pitch 範圍 """
        InfoMgr.minPitch = self._noteStrToPitch(
            InfoMgr.limit[LimitData.scale.value]["min"])
        InfoMgr.maxPitch = self._noteStrToPitch(
            InfoMgr.limit[LimitData.scale.value]["max"])

    def _noteStrToPitch(self, noteStr: str):
        """ 純文字音符變成 pitch
        :param noteStr: 純文字音符，例如 7'
        """
        up: int = noteStr.count("'")
        down: int = noteStr.count(",")
        note: int = noteStr[: 1]
        return ScalePitch[ScalePitch.scaleName.value[int(note) - 1]].value + (up - down) * ScalePitch.step.value

    def _repeatSentence(self, sentenceIndex: int):
        """ 重複句子 -> bool (True 有重複，False 沒有重複需要生成新句子) """
        repeatCount = int(
            InfoMgr.weights[WeightsData.repeatCount.value].getRandKey())
        # 重複很多句子
        if repeatCount <= self.repeatNum and repeatCount > 0:
            for index in range(repeatCount):
                nowIndex = sentenceIndex + index
                sentenceObject = SentenceObject(pos=nowIndex)
                sentenceObject.repeatFrom(
                    InfoMgr.music[nowIndex - repeatCount])
                InfoMgr.music[nowIndex] = sentenceObject
            self.repeatNum = 0
            return True
        else:
            return False

    def _createMIDI(self):
        MyMIDI = MIDIFile(1)
        track = 0   # the only track
        channel = 0
        time = 0    # start at the beginning
        MyMIDI.addTrackName(track, time, "track" + str(track))
        MyMIDI.addTempo(track, time, int(
            InfoMgr.limit[LimitData.speed.value]["bpm"]))
        MyMIDI.addProgramChange(0, 0, 0, int(
            InfoMgr.limit[LimitData.program.value]["value"]))
        for sentenceIndex in InfoMgr.music:
            for note in InfoMgr.music[sentenceIndex].notes:
                MyMIDI.addNote(track=track, channel=channel, pitch=note.pitch,
                               time=time, duration=note.duration, volume=note.volume)
                time += note.duration
            time += 1
        with open(datetime.now().strftime('%Y%m%d-%H%M%S') + ".mid", "wb") as midi_file:
            MyMIDI.writeFile(midi_file)


a = Main()
